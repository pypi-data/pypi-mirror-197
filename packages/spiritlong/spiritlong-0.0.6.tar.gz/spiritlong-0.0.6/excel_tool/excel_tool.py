## 
# Copyright (c) 2023 Chongqing Spiritlong Technology Co., Ltd.
# All rights reserved.
# 
# @author	arthuryang
# @brief	office工具集，目前包括对xlsx和docx的支持。
##

import	openpyxl
import	re

## 打开xlsx文件
def open_xlsx(filename):
	try:
		book	= openpyxl.load_workbook(filename)
		return book
	except Exception as ex:
		print(str(ex))
		exit

## excel读取出None的情况要转换成空字符串
def cell_string(sheet, i, j):
	value	= sheet.cell(i, j).value
	if not value:
		value	= ""
	return str(value)

## 填充excel单元格，可以指定格式
def cell_fill(sheet, i, j, value, style={
					'font'		: None,
					'fill'		: None, 
					'border'	: None, 
					'alignment'	: None, 
					'number_format'	: None,
				}):
	cell		= sheet.cell(row=i, column=j)
	cell.value	= value
	
	if 'font' in style and style['font']:
		cell.font		= style['font']
	if 'fill' in style and style['fill']:
		cell.fill		= style['fill']
	if 'border' in style and style['border']:
		cell.border		= style['border']
	if 'alignment' in style and style['alignment']:
		cell.alignment		= style['alignment']
	if 'number_format' in style and style['number_format']:
		cell.number_format	= style['number_format']

## 自动调整指定列的宽度
def adjust_column_width(sheet, column, max_width=100):
	width	= 1
	for row in range(1, sheet.max_row+1):
		# 对于中文字符和非中文字符单独计算宽度
		s	= cell_string(sheet, row, column)
		w	= 0
		for c in s:
			w	+= (1 if ord(c)<128 else 2)
		width	= max(width, w+1)
	if width>max_width:
		# 列宽得有最大限制，不然无法弄
		width	= max_width
	sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width	= width

## 自动调整所有列的宽度
def adjust_all_column_width(sheet, max_width=100):
	for column in range(1, sheet.max_column+1):
		adjust_column_width(sheet, column, max_width)

## 将'A1'转换为（1,1)
def cell_coordinate(cell_string):
	column_letters, row	= openpyxl.utils.cell.coordinate_from_string(cell_string)
	column			= openpyxl.utils.cell.column_index_from_string(column_letters)
	return (row, column)

## 将（1，1）转换为'A1'
def cell_code(row, column):
	return f"{openpyxl.utils.cell.get_column_letter(column)}{row}"

# excel单元格格式：Calibri字体，11号，黑色
font_default	= openpyxl.styles.Font(
	name		= 'Calibri',
	size		= 11,
	bold		= False,
	vertAlign	= None,
	underline	= 'none',
	strike		= False,
	color		= 'FF000000',
)

# excel单元格格式：Calibri字体，11号，白色
font_title	= openpyxl.styles.Font(
	name		= 'Calibri',
	size		= 11,
	bold		= False,
	vertAlign	= None,
	underline	= 'none',
	strike		= False,
	color		= 'FFFFFFFF',
)

# excel单元格格式：Calibri字体，11号，红色
font_red	= openpyxl.styles.Font(
	name		= 'Calibri',
	size		= 11,
	bold		= False,
	vertAlign	= None,
	underline	= 'none',
	strike		= False,
	color		= 'FFFF0000',
)

# excel单元格标题格式：蓝色填充
fill_title	= openpyxl.styles.PatternFill(
	fill_type	= 'solid',
	fgColor		= 'FF4169E1',
)

# excel单元格格式：居中对齐
alignment_center	= openpyxl.styles.Alignment(
	horizontal	= 'center',
	vertical	= 'center',
	wrap_text	= False,
)

# excel单元格格式：左对齐
alignment_left	= openpyxl.styles.Alignment(
	horizontal	= 'left',
	vertical	= 'center',
	wrap_text	= False,
)

# excel单元格格式：右对齐
alignment_right	= openpyxl.styles.Alignment(
	horizontal	= 'right',
	vertical	= 'center',
	wrap_text	= False,
)

################ 单元格样式：字体 ################
# 默认黑色11号Calibri
def font(name='Calibri', size=11, color='FF000000', bold=False):
	return openpyxl.styles.Font(
		name		= name,
		size		= size,
		bold		= bold,
		vertAlign	= None,
		underline	= 'none',
		strike		= False,
		color		= color,
	)

################ 单元格样式：填充 ################
# 默认白色
def fill_with_color(color='FFFFFFFF'):
	return openpyxl.styles.PatternFill(
		fill_type	= 'solid',
		fgColor		= color,
	)

################ 单元格样式：对齐 ################
# 居中对齐
alignment_center	= openpyxl.styles.Alignment(
	horizontal	= 'center',
	vertical	= 'center',
	wrap_text	= False,
)

# 左对齐
alignment_left	= openpyxl.styles.Alignment(
	horizontal	= 'left',
	vertical	= 'center',
	wrap_text	= False,
)

# 右对齐
alignment_right	= openpyxl.styles.Alignment(
	horizontal	= 'right',
	vertical	= 'center',
	wrap_text	= False,
)

# 调试/测试代码
if __name__ == '__main__':
	pass